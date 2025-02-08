import json
from openpyxl import load_workbook
import re
from datetime import datetime

# 엑셀 파일 경로 설정
file_path = r"D:\컬쳐마케팅컴퍼니\doc\홈페이지개편_포트폴리오 구분(국문,영문)_ver.0.1.xlsx"  # 실제 엑셀 파일 경로로 변경

# 엑셀 파일 로드
wb = load_workbook(file_path)
sheet_names = wb.sheetnames  # 모든 시트 이름 가져오기

# 날짜 정규화 함수
def normalize_date(date_str, global_year=None, ref_date=None, language='korean'):
    print(f"  {date_str}", end=' / ')
    print(f"ref_date = {ref_date}", end=' / ')
    """다양한 형식의 날짜 문자열을 YYYY-MM-DD 형식으로 표준화"""
    # 괄호와 요일 정보 제거
    date_str = re.sub(r'\([^)]*\)', '', date_str).strip()

    # 연도, 월, 일 추출
    parts = re.split(r'[./-]', date_str)
    parts = [p.strip() for p in parts if p.strip()]

    year, month, day = None, None, None

    if len(parts) == 3:
        year, month, day = parts
    elif len(parts) == 2:
        month, day = parts
    elif len(parts) == 1:
        day = parts[0]

    print(f"2. YYYY-MM-DD: {year}-{month}-{day}", end=" / ")


    # 연도와 월 추출 시도
    year_month_match = re.search(r'(\d{4})년?\s*(\d{1,2})월?', date_str)
    if year_month_match:
        year, month = year_month_match.groups()
        print(f"3. YYYY-MM-DD: {year}-{month}-{day} result: {year}-{month.zfill(2)}")
        return f"{year}-{month.zfill(2)}"

    # 연도만 있는 경우
    year_match = re.search(r'\d{4}', date_str)
    if month is None and day is None and year_match:
        print(f"5. YYYY-MM-DD: {year}-{month}-{day} result: {year_match.group()}")
        return year

    # 연도 처리
    if year is None:
        if global_year:
            year = global_year
        elif ref_date:
            year = ref_date[:4]
        else:
            year = str(datetime.now().year)

    # 월 처리
    if month is None and ref_date:
        month = ref_date[5:7]

    # 연도가 두 자리인 경우 네 자리로 확장
    if len(year) == 2:
        year = '20' + year

    # 월과 일이 한 자리인 경우 두 자리로 패딩
    if month:
        month = month.zfill(2)
    if day:
        day = day.zfill(2)

    try:
        result = ''
        # 유효한 날짜인지 확인
        if month and day:
            datetime.strptime(f"{year}-{month}-{day}", "%Y-%m-%d")
            result = f"{year}-{month}-{day}"
        elif month:
            result = f"{year}-{month}"  # 일자가 없는 경우
        else:
            result = f"{year}"  # 월과 일자가 없는 경우
        print(f"result: {result}")
        return result
    except ValueError:
        print(f"ValueError: {date_str}")
        return date_str  # 유효하지 않은 날짜인 경우 원본 반환

# 개요 파싱 함수
def parse_outline(text, language='korean', global_year=None, skip_dates=False):
    outline_data = {}
    sections = re.split(r'•', text)[1:]  # 필드 단위 분리

    for section in sections:
        if ':' not in section:
            continue

        field_part, *content_part = section.split(':', 1)
        field_name = field_part.strip()
        content = content_part[0].strip() if content_part else ''

        # 언어별 필드명 매핑
        field_map = {
            'korean': {'일시': 'date', '장소': 'location', '대상': 'target', '주최/주관': 'organizer', '주제/내용': 'topic'},
            'english': {'Time': 'date','Period': 'date', 'Place': 'location', 'Target': 'target', 'Host/Supervisor': 'organizer', 'Subject/Content': 'topic'}
        }

        # 내용 처리
        if content.startswith('-'):
            items = [line.strip('- ').strip() for line in content.split('\n') if line.strip().startswith('-')]
            processed_content = items
        else:
            processed_content = content.replace('\n', ', ')

        # 날짜 필드 처리
        if field_name in ['일시', 'Time', 'Period']:
            outline_data['dateText'] = content  # 원본 날짜 보존
            dates = [d.strip() for d in processed_content.split('~')]
            print(dates)
            if len(dates) > 1:
                start_date = normalize_date(dates[0], global_year)
                end_date = normalize_date(dates[1], global_year, ref_date=start_date)
                outline_data['startDate'] = start_date
                outline_data['endDate'] = end_date
            else:
                outline_data['startDate'] = normalize_date(dates[0])
                outline_data['endDate'] = outline_data['startDate']

        # 기본 필드 매핑
        elif field_name in field_map[language]:
            mapped_field = field_map[language][field_name]
            outline_data[mapped_field] = processed_content

        # 추가 필드 처리
        else:
            outline_data[field_name] = processed_content

    return outline_data


# 데이터 추출 및 JSON 변환
all_data = []

for idx, sheet_name in enumerate(sheet_names):
    sheet = wb[sheet_name]

    # 시트별 카테고리 설정
    if idx == 0:
        category = "pubGov"
    elif idx == 1:
        category = "performance"
    elif idx == 2:
        category = "event"
    else:
        category = "unknown"

    for row in sheet.iter_rows(min_row=8, values_only=True):
        if not any(row):
            continue

        title = row[2]
        if not title:  # title이 없으면 다음 시트로 넘어감
            break

        # 한국어 개요 파싱 (날짜 계산)
        outline = parse_outline(row[5], 'korean', str(row[1])) if row[5] else None

        # 영어 개요 파싱 (날짜 계산 생략)
        outline_en = parse_outline(row[6], 'english', skip_dates=True) if row[6] else None

        # 영어 개요에 한국어 날짜 복사
        if outline and outline_en:
            outline_en['startDate'] = outline.get('startDate')
            outline_en['endDate'] = outline.get('endDate')

        entry = {
            "category": category,
            "year": row[1],
            "title": title,
            "title_en": row[3],
            "institution": row[4],
            "outline": outline,
            "outline_en": outline_en
        }

        all_data.append(entry)

# JSON 저장
output_file = "output.json"
with open(output_file, "w", encoding="utf-8") as f:
    json.dump(all_data, f, ensure_ascii=False, indent=4)

print(f"데이터가 JSON으로 저장되었습니다: {output_file}")
